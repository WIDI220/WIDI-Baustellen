from http.server import BaseHTTPRequestHandler
import json, io, base64, re, csv
from collections import Counter

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    HAS_OPX = True
except ImportError:
    HAS_OPX = False

try:
    import pandas as pd
    HAS_PD = True
except ImportError:
    HAS_PD = False

NORMBEZ = ["Airhockeytisch","Aktenvernichter","Anschlusskabel","Apple TV","Aquariumzubehör","Babyphone","Barcode / QR-Codescanner","Batterietester","Beamer","Beistellmodul (Telefon)","Beschriftungsgerät","Bettbeleuchtung","Bildimporter","Blaulichtlampe","Blu-ray / DVD Player","Bohrmaschine","Bratpfanne","Brutschrank","Bügelbrett","Bügeleisen","CD / Radio / Kassettenrekorder","Crêpes-Eisen","Dartscheibe (elektrisch)","Deckenfluter","Dekobeleuchtung","Desinfektionsmittelmischgerät","Digitaler Fotorahmen","Diktiergerät","Docking Station","Dosieranlage","Dreiwalzwerk","Drucker","Duftlampe","Dunstabzugshaube","Durchlauferhitzer","Eierkocher","Einkochautomat","Einschaltstrombegrenzer","Elektroherd","Elektronische Wetterstation","Entsafter","Etikettendrucker","Falschgelddetektor","Faltmaschine","Fango-Ofen","Faxgerät","Fernseher","Flaschenwärmer","Fleischwolf (elektrisch)","Fliegenfalle (elektrisch)","Fritteuse","Fräsmaschine","Fußschalter","Fußwärmer","Fön","Gefrierschrank","Geldzähler","Globus (elektrisch)","Glühweinkocher","Graviergerät","Grill (elektrisch)","Haarschneidemaschine","Haartrockner","Handlampe","Handmixer","Handnotleuchte","Handstaubsauger","Headset","Heizbad","Heizdecke","Heizkissen","Heizlüfter","Heizung (elektrisch)","Heißgeräteanschlussleitung","Heißklebepistole","Heißluftfön","Infrarotlampe","Kabeltrommel","Kaffeemaschine","Kaffeemühle","Kaffeevollautomat","Kalender (digital)","Kaltgeräteanschlussleitung","Kamin (elektrisch)","Kartenlesegerät","Kartenzahlungsgerät","Keyboard","Kiesbad","Kochplatte","Konferenztelefon","Konvektor","Körperwaage","Küchenmaschine","Küchenwaage","Kühl- und Gefrierkombination","Kühlschrank","Kühltruhe","LED Treiber","Labeldrucker","Ladegerät / Ladestation","Laminiergerät","Lasergerät","Lautsprecher","Lesehilfe","Lichterbogen","Lichterkette","Liege (elektrisch)","Lockenstab","Luft-Kompressor","Luftfilteranlage","Lötkolben","Lüfter","Massagegerät (Infrarot)","Mehrfachsteckdose","Mikrofon","Mikrowelle","Milchaufschäumer","Milchwaage","Mischpult","Modem","Monitor","Multifunktionsgerät (Drucker&Scanner)","Nachtlicht","Nähmaschine","Overhead Projektor","PC","PC (mobil)","PC Arbeitsplatz","Parafinbad","Parkscheinentwerter","Partytopf","Patchmaschine","Pendelleuchte","Popcornmaschine","Pumpe (elektrisch)","Rasierer","Receiver","Rechenmaschine","Router/Switch","Sandwichmaker","Scanner","Scheinwerfer","Schleifmaschine","Schließfach","Schneidemaschine","Schokoladenbrunnen","Schranklampe","Schreibmaschine (elektrisch)","Schreibtischlampe","Schwarzlichtlampe","Sessel (elektrisch)","Spielekonsole","Sportgerät (elektrisch)","Spülmaschine","Stabmixer","Standbohrmaschine","Standlampe","Standmixer","Staub/Nasssauger","Steckdose mit Schalter","Steckernetzteil","Stift","Stromverteiler 400V","Säge","Tacker (elektrisch)","Tafelsteuerung","Tauchpumpe","Teekocher","Teigknetmaschine","Tellerwärmer","Temperaturfühler","Tiefkühltruhe","Tisch (elektrisch)","Tischfräse","Toaster","Topf (elektrisch)","Trockner","USV (Unterbrechungsfreie Stromversorgung)","Uhr","Ultraschallreinigungsgerät","Untertischgerät","VGA Switch","Ventilator","Verlängerung 230V","Verlängerung 400V","Vernebler","Verstärker","Videokamera","Videokonferenzanlage","Videorekorder","Waffeleisen","Waschmaschine","Wasserbett","Wasserbrunnen","Wasserkocher","Wasserspender","Wecker","Whiteboard","Wok (elektrisch)","Wärmematte","Wärmeplatte","Wärmeschrank","Wärmestrahler","Wärmewagen","Zahnbürste (elektrisch)","Zeichentisch","Zeitschaltuhr (elektrisch)","Zigarettenstopfgerät","Zimmerantenne","iPad","mobiles Klimagerät","Überspannungsschutz Feinschutz Typ 3"]

TABELLE3_HEADERS = ['Raum-ID','Raumbez.','Kostenstelle','Kostenstellenbez.','Liegenschafts-ID','Gesellschaft','Kundenbez.','WIDI Kunden Nr']

COLS = ['OBJEKTTYP','ID','BEZEICHNUNG','TYP','SERIENNUMMER','HERSTELLER','STATUS',
        'INTERVALL (MONATE)','LETZTE PRÜFUNG','LETZTER PRÜFER','STATUS TERMIN',
        'ERGEBNIS DER LETZTEN PRÜFUNG','NÄCHSTE PRÜFUNG','ABTEILUNG','KOSTENSTELLE',
        'DOKUMENTE','PRÜFSEQUENZ','AKTIV','ERFASST AM','GEÄNDERT AM','ERFASST VON',
        'BEMERKUNG','KUNDENBEZEICHNUNG','KUNDEN-ID','LIEGENSCHAFT','LIEGENSCHAFTS-ID',
        'GEBÄUDE','GEBÄUDE-ID','EBENE','EBENEN-ID','RAUM','RAUM-ID','Neu?']

COL_WIDTHS = [12,7,42.4,18.6,25.4,16.1,10.4,20.9,18.1,17.4,16.3,31.1,19.3,23.3,
              15.1,14.6,15.0,7.7,13.0,15.1,14.0,13.7,76.0,19.4,15.1,18.4,10.9,
              13.3,8.3,12.1,8.1,10.6,13.1]

YELLOW_COLS = {3, 14}

def fuzzy_score(a, b):
    s,t = a.lower().strip(), b.lower().strip()
    if s==t: return 1.0
    if t in s or s in t: return 0.85
    matches = sum(1 for c in min([s,t],key=len) if c in max([s,t],key=len))
    return matches/max(len(s),len(t))

def find_best_norm(bez):
    if not bez: return bez
    bez = bez.strip()
    exact = next((n for n in NORMBEZ if n.lower()==bez.lower()), None)
    if exact: return exact
    best_score, best_match = 0, bez
    for norm in NORMBEZ:
        score = fuzzy_score(bez, norm)
        if score > best_score: best_score, best_match = score, norm
    return best_match if best_score > 0.75 else bez

def format_date(val):
    if not val: return None
    val = str(val).strip()
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$', val)
    if m:
        year = f"20{m.group(3)}" if len(m.group(3))==2 else m.group(3)
        return f"{year}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    return val if val not in ('nan','None','') else None

def extract_raum(abteilung):
    if not abteilung: return ''
    m = re.search(r'\.(R\d{3,4})', str(abteilung))
    return m.group(1) if m else ''

def shorten_kunde(val):
    if not val: return val
    val = re.sub(r'^\d+\s*-\s*', '', str(val).strip())
    return re.sub(r',.*$', '', val).strip()

def parse_csv(text):
    text = text.replace('\r\n','\n').replace('\r','\n')
    lines = [l for l in text.split('\n') if l.strip()]
    if len(lines) < 2: return []
    sep = ';' if lines[0].count(';') > lines[0].count(',') else ','
    reader = csv.DictReader(lines, delimiter=sep)
    return [dict(r) for r in reader]

def process_rows(rows):
    aenderungen = []
    processed = []
    for row in rows:
        r = {k: (str(v).strip() if v is not None else '') for k,v in row.items()}
        id_ = r.get('ID','').strip()

        # 1. Bezeichnung
        bez = r.get('BEZEICHNUNG','')
        if bez:
            neu = find_best_norm(bez)
            if neu != bez:
                aenderungen.append({'id':id_,'feld':'BEZEICHNUNG','vorher':bez,'nachher':neu,'typ':'auto'})
                r['BEZEICHNUNG'] = neu

        # 2. Datum
        for df in ['LETZTE PRÜFUNG','NÄCHSTE PRÜFUNG','ERFASST AM','GEÄNDERT AM']:
            v = r.get(df,'')
            if v and '.' in v:
                neu = format_date(v)
                if neu: r[df] = neu

        # 3. RAUM
        abt = r.get('ABTEILUNG','')
        raum = extract_raum(abt)
        if raum and not r.get('RAUM'):
            r['RAUM'] = raum
            aenderungen.append({'id':id_,'feld':'RAUM','vorher':'','nachher':raum,'typ':'auto'})

        # 4. Kundenbezeichnung
        k = r.get('KUNDENBEZEICHNUNG','')
        if re.match(r'^\d+\s*-\s*', k):
            kurz = shorten_kunde(k)
            r['KUNDENBEZEICHNUNG'] = kurz
            aenderungen.append({'id':id_,'feld':'KUNDENBEZEICHNUNG','vorher':k,'nachher':kurz,'typ':'auto'})

        # 5. Neu?
        bem = r.get('BEMERKUNG','').strip().lower()
        if bem in ['neu','neru'] or bem.startswith('neu,') or bem.startswith('neu '):
            r['Neu?'] = 'NEU'

        processed.append(r)
    return processed, aenderungen

def build_excel(rows):
    wb = Workbook()

    # Tabelle6 - Pivot
    ws6 = wb.active
    ws6.title = 'Tabelle6'
    ws6.column_dimensions['A'].width = 60
    ws6.column_dimensions['B'].width = 30
    kunde_counts = Counter()
    for r in rows:
        k = r.get('KUNDENBEZEICHNUNG','').strip()
        if k: kunde_counts[k] += 1
    ws6['A3'] = 'Zeilenbeschriftungen'
    ws6['B3'] = 'Anzahl von KUNDENBEZEICHNUNG'
    ws6['A3'].font = Font(bold=True, size=11)
    ws6['B3'].font = Font(bold=True, size=11)
    ri = 4; total = 0
    for kunde, count in sorted(kunde_counts.items()):
        ws6.cell(ri,1,kunde); ws6.cell(ri,2,count); total += count; ri += 1
    ws6.cell(ri,1,'Gesamtergebnis').font = Font(bold=True)
    ws6.cell(ri,2,total).font = Font(bold=True)

    # Tabelle1 - Hauptdaten mit GRÜNEM Tabellenstil
    ws1 = wb.create_sheet('Tabelle1')
    yellow_fill = PatternFill('solid', fgColor='FFFF00')

    for ci, col_name in enumerate(COLS, 1):
        cell = ws1.cell(1, ci, col_name)
        cell.font = Font(size=11)
        if ci in YELLOW_COLS: cell.fill = yellow_fill

    for ci, width in enumerate(COL_WIDTHS, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = width

    for ri2, r in enumerate(rows, 2):
        for ci, col_name in enumerate(COLS, 1):
            val = r.get(col_name, '')
            if val in ('','nan','None'): val = None
            ws1.cell(ri2, ci, val).font = Font(size=11)

    # Grüner Tabellenstil TableStyleMedium14
    last_row = len(rows) + 1
    last_col = get_column_letter(len(COLS))
    tbl = Table(displayName="Tabelle2", ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium14",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws1.add_table(tbl)

    # Tabelle2 - Normbezeichnungen
    ws2 = wb.create_sheet('Tabelle2')
    ws2.column_dimensions['A'].width = 40
    ws2['A1'] = 'Anlagenbezeichnung'
    ws2['A1'].font = Font(size=10)
    for i, bez in enumerate(NORMBEZ, 2):
        ws2.cell(i,1,bez).font = Font(size=10)
    t2 = Table(displayName="Tabelle3", ref=f"A1:A{len(NORMBEZ)+1}")
    t2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium14", showRowStripes=True)
    ws2.add_table(t2)

    # Tabelle3 - Raum-Mapping Header
    ws3 = wb.create_sheet('Tabelle3')
    gray_fill = PatternFill('solid', fgColor='AAAAAA')
    for ci, h in enumerate(TABELLE3_HEADERS, 1):
        cell = ws3.cell(1, ci, h)
        cell.fill = gray_fill
        cell.font = Font(bold=True, size=10)

    # Tabelle4 leer
    wb.create_sheet('Tabelle4')
    wb._sheets = [ws6, ws1, ws2, ws3, wb['Tabelle4']]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Methods','POST,OPTIONS')
        self.send_header('Access-Control-Allow-Headers','Content-Type')
        self.end_headers()

    def do_POST(self):
        if not HAS_OPX:
            self._json(500, {'error':'openpyxl not installed'})
            return
        length = int(self.headers.get('Content-Length',0))
        body = self.rfile.read(length)
        try:
            data = json.loads(body)
            csv_text = data.get('csv','')
            rows = parse_csv(csv_text)
            processed, aenderungen = process_rows(rows)
            excel_bytes = build_excel(processed)
            excel_b64 = base64.b64encode(excel_bytes).decode()
            self._json(200, {
                'success': True,
                'excel_b64': excel_b64,
                'aenderungen': aenderungen[:500],
                'stats': {
                    'gesamt': len(processed),
                    'auto': len([a for a in aenderungen if a['typ']=='auto']),
                    'bezeichnung': len([a for a in aenderungen if a['feld']=='BEZEICHNUNG']),
                    'raum': len([a for a in aenderungen if a['feld']=='RAUM']),
                    'kunde': len([a for a in aenderungen if a['feld']=='KUNDENBEZEICHNUNG']),
                }
            })
        except Exception as e:
            self._json(500, {'success':False,'error':str(e)})

    def _json(self, code, data):
        body = json.dumps(data).encode()
        self.send_response(code)
        self.send_header('Content-Type','application/json')
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Content-Length',str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *args): pass
